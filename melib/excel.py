"""
    **excel.py**

    Methods that are used in ``melib`` to read, edit and generate ``excel`` files

"""
# excel
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles import Font, colors, Color
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
#
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import math
import sys
# from xt import thesame

COLSKIP=0  # Empty columns on the left of the Mark column (default=0)
DEBUG = True
MAXTABLECOLUMNS=20  # Maximum number of data columns in a table
TESTMODULE=""
# TESTMODULE="fontcolor"
# TESTMODULE="w2row"
# TESTMODULE="vit"

def similar(x, y, eps=0.001):
    try:
        x=float(x)
    except:
        x=0
    try:
        y=float(y)
    except:
        y=0
    diff=abs(x-y)
    xysum=abs(x+y)
    if xysum==0:
        if abs(x)>eps:
            return False
        else:
            return True
    else:
        r=diff/xysum
        if r>eps:
            return False
        else:
            return True


# There should not be any blank rows until the end of the table in either file
# There should be a label in the second column of each row
# No label means the end of the table
def sametable(table_num, xf1, xf2, tablecol=0, skip={}):
    irow=xf1.findtable(table_num)+1
    while irow<xf1.maxrow():
        if xf1.cell(irow,3+COLSKIP)==None:
            return ""  # The table ended but no difference was observed
        parname=xf1.cell(irow, 3)
        if parname in skip.keys():
            tol=skip[parname]            
        else:
            tol=0.001
        v1=xf1.cell(irow, tablecol+6+COLSKIP)
        v2=xf2.cell(irow, tablecol+6+COLSKIP)
        cell=xf1.ws.cell(row=irow, column=tablecol+6+COLSKIP)
        color_index=cell.fill.start_color.index
        if type(color_index)==int:
        # Protection against stupid Excel feature.  The colors selected from
        # "theme colors" menu in Excel are integers. There is no color match
        # between my "Excel theme color" indices and what openpyxl thinks they are.
        # I think it is always a string when you pick the color from the color wheel
        # Therefore, I will not accept integers and force selection from the wheel.
            # import sys
            # sys.exit("'%s'(%d,%d) : You must select from the color wheel not from 'standard colors'"%(cell.value,row,col))
            colored=True
        else:
            colored=int(color_index,16)>0

        if not colored and (not similar(v1,v2,eps=tol) and not v2==None):
            s="\nirow=%d(%s:%s)MISMATCH "%(irow,xf1.cell(irow,5),xf2.cell(irow,5))
            print(s)
            print(v1, end="")
            print(" and ", end="")
            print(v2)
            return s
        irow+=1
    return ""


#colortable={"RED":colors.RED, "BLACK":colors.BLACK, "GREEN":colors.GREEN,
#"DARKBLUE":colors.DARKBLUE, "DARKRED":colors.DARKRED}
colortable={'RED': '00FF0000',
 'BLACK': '00000000',
 'GREEN': '0000FF00',
 'DARKBLUE':'000000FF',
 'DARKRED': '00000000',
 'PURPLE':  '800080',
 'BLUE':'0000FF',
 'GRAY':'808080'}

def column_index(s):
    col = column_index_from_string(s)
    return col

# data_only: This is a confusing parameter.  You can either have the
# formula or the cached value of the last evaluation. If you alter a file
# with formulae then you must pass it through some kind of application such
# as Excel if you want to get the value of the formulae.
# Otherwise, openpyxl does not return the value but returns "None"
# When "data_only", the openpyxl reads the value stored the last time
# Excel read the sheet. If the file is direct from another openpyxl program,
# then it will read is as "None".
# This is is why I have the additional lines to process the formulae.
class Xcel:
    lastrow=0
    Debug=False
    Xcelfilename=""
    def __init__(self, filename=None, sheetname="", data_only=True):
        import os
        self.fdic={}  # Just in case I want to add a dictionary of funcs
        if filename==None:
            self.wb=Workbook()
            self.ws=self.wb.active
            if sheetname=="":
                sheetname="data"
            self.ws.title=sheetname
        elif os.path.isfile(filename):
            if (self.openfile(filename, sheetname, data_only=data_only)):
                self.Xcelfilename=filename
            else:
                self.Xcelfilename=""
        else:
            sys.exit("Xcel('%s') cannot be found."%filename)



    def filename(self):
        return self.Xcelfilename

    def newsheet(self, sheetname):
        self.ws=self.wb.create_sheet(title=sheetname)

    def sheet(self, sheetname=None):
        if not sheetname==None:
            try:
                self.ws=self.wb[sheetname]
            except:
                return None
        return self.ws.title

    def debug(self, debug=True):
        self.Debug=debug

    def protect(self, password):
#        self.ws.protection.set_password(password)
#        self.ws.protection.SheetProtection(insertRows=False, password=password)
        self.ws.protection.enable()
        self.ws.protection.password=password
# SOURCE = https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.protection.html

    def cellprotect(self, row, column, protect):
        cello=self.ws.cell(row=row, column=column)
        cello.protection=Protection(locked=protect, hidden=False)

    # Column Width, can be called by c="D" or c=4
    def columnw(self,c,w):
        if type(c)==int:
            c=chr(ord('A')+c-1)
        self.ws.column_dimensions[c].width=w

    def setws(self, ws):
        self.ws=ws
    def maxrow(self):
        return self.ws.max_row
    def getrow(self, rowname):
        ws=self.ws
        k=0
        for row in ws.rows:
            k+=1
            if len(row)>3 and rowname==row[2].value:
                return k
        return None
    def vir(self, rowname, col=6, value=None, sheet=None):  # vir: valueinrow
        self.sheet(sheetname=sheet)
        row=self.getrow(rowname)
        if row==None:
            return None
        return self.cell(row, col, value)
    def virs(self, rowname, col=6, sheet=None):
        self.sheet(sheetname=sheet)
        row=self.getrow(rowname)
        if row==None:
            return None
        return self.cell(row, col-1)
        
    def params(self, rownames):
        v=[]
        for rowname in rownames:
            v.append(self.vir(rowname))
        return v
    def findrow(self, rowname, column=2):
        k=0
        for row in self.ws.rows:
            k+=1
            if len(row)>3 and rowname==row[column].value:
                return k
        return None
    def rowlabel(self, jrow):
            return self.cell(jrow, 3+COLSKIP)

    def findtable(self, table_num):
        k=0
        for row in self.ws.rows:
            k+=1
            if row[2].value=="Table" and row[3].value==table_num:
                return k
        return None

    def findtablecolumn(self, table_num, colname, startcolumn=6):
        tablerow=self.findtable(table_num)
        j=startcolumn
        scol=self.cell(tablerow+1, j)
        while (scol!=None):
            if scol==colname:
                return j
            j+=1
            scol=self.cell(tablerow+1, j)  # The heading for column j
        return 0

    def findtablerow(self, table_num, name_wanted):
        i = self.findtable(table_num)+1
        if name_wanted=='FIRSTDATAROW':
            return i+4
        nrows=0
        rowname=self.cell(i, 3)
        while (rowname!=None or nrows<4):
            if rowname==name_wanted:
                return i
            i+=1
            nrows+=1
            rowname=self.cell(i, 3)
    def mdtable(self, table_num, md, startcolumn=6, showlabel=False):
        irow=self.findtablerow(table_num, "FIRSTDATAROW")
        heading_row=irow-4
        x=self.cell(heading_row, startcolumn)
        n=0
        headings=[]
        while x!=None:
#            headings.append(self.cell(irow-4,startcolumn+n))
            headings.append(x)
            n+=1
            x=self.cell(heading_row,startcolumn+n)
        if showlabel:
            md.write("|Label")
        md.write("|Description|")
        for i in range(0,n):
            md.write("%s|"%headings[i])
        md.write("\n")
        if showlabel:
            md.write("|:---")
        md.write("|:---|")
        for i in range(0,n):
            md.write(":--:|")
        md.write("\n")
        x=self.cell(irow, startcolumn-1)
        while x!=None:
            if showlabel:
                label=self.cell(irow, startcolumn-3)
                if label==None:
                    md.write("| ")
                else:
                    md.write("|%s"%label)
            md.write("|%s"%x)
            for j in range(0,n):
                y=self.cell(irow, startcolumn+j)
                if y==None:
                    md.write("| ")
                else:
                    md.write("|%s"%y)
            md.write("|\n")
            irow+=1
            x=self.cell(irow, startcolumn-1)

    # VIT()
    # if "value"=None:
    #       return the value in the cell (rowname, colname)
    # else:
    #       set the value of the cell to "value"
    def vit(self, table_num, rowname, colname, startcolumn=6, value=None, sheet=None):  # valueintable
        self.sheet(sheetname=sheet)
        tablerow=self.findtable(table_num)
        k=tablerow
        nrows=0
#        print("Found table : k=%d"%k)
        srow=self.cell(k, 3)
        while (srow!=None or nrows<4):
            if srow==rowname:
#                print("Found row: %s.  Now search for %s."%(srow, colname))
                j=startcolumn
                if colname=="":
                    return self.cell(k, j, value=value)
                scol=self.cell(tablerow+1, j)
                while (scol!=None):
                    # if srow==rowname and scol==colname:
                    if scol==colname:
                        return self.cell(k, j, value=value)
                    j+=1
                    scol=self.cell(tablerow+1, j)
#            else:
#                print("Looking for %s, skip %s"%(rowname, srow))
            k+=1
            nrows+=1
            srow=self.cell(k, 3)
    def vits(self, table_num, rowname, startcolumn=6, sheet=None):
        self.sheet(sheetname=sheet)
        tablerow=self.findtable(table_num)
        k=tablerow
#        print("Found table : k=%d"%k)
        srow=self.cell(k, 3)
        while (srow!=None or k<4):
            if srow==rowname:
                return self.cell(k, startcolumn-1)
            k+=1
            srow=self.cell(k, 3)
        return None
        
    # Read (or write) the values in a table row.  The default is read.
    # The array "values" is used to pass the values back
    def valuesintable(self, table_num, rowname, values, startcolumn=6, readonly=True):
        tablerow=self.findtable(table_num)
        k=tablerow
        srow=self.cell(k, 3)
        jmax=min(MAXTABLECOLUMNS, len(values))
        while (srow!=None):
            if srow==rowname:
                j=0
                while j<jmax:
                    if readonly:
                        values[j]=self.cell(k, j+6)
                    else:
                        self.cell(k, j+6, value=values[j])
                    j+=1
            k+=1
            srow=self.cell(k, 3)

    # Read "n" values from "rowname" in thetable
    def rnvit(self, table_num, rowname, cols, startcolumn=6):
        n = len(cols)
        v = np.zeros(n)
        i = 0
        for s in cols:
            try:
                v[i]=self.vit(table_num, rowname, s, startcolumn=startcolumn)
            except:
                v[i]=0.0
            i+=1
        return v

    # Read n numbers from the Table column COLUMNNAME starting from ROWNAME
    # If the number is not a number, it forces it to be a number
    def rnvitc(self, table_num, rowname, columnname, n):
        v=np.zeros(n)
        irow=self.findtablerow(table_num,rowname)
        jcol=self.findtablecolumn(table_num,columnname)
        for j in range(0,n):
            c=self.cell(irow+j, jcol)
            try:
                v[j]=float(c)
            except:
                v[j]=0
        return v

    # Write "n" values to "rowname" in thetable
    def wnvit(self, table_num, rowname, cols, v, startcolumn=6):
        n = len(cols)
        i = 0
        for s in cols:
            self.vit(table_num, rowname, s, startcolumn=startcolumn, value=v[i])
            i+=1
    # Write to table row (len(v) values contiguously)
    def w2trow(self, table_num, rowname, v, startcolumn=6):
        irow=self.findtablerow(table_num, rowname)
        self.w2row(irow, startcolumn, v)

    def w2row(self, irow, startcolumn, v, color="BLACK", font=None, align=None):
        j=0
        for x in v:
            self.cell(irow, j+startcolumn, color=color, value=v[j],font=font, align=align)
            j+=1
    # Write an array on a row
    # def wrow(self, row, col, a, color="BLACK", fontsize=None):
    #     n=len(a)
    #     for j in range(0, n):
    #         self.cell(row, col+j, value=a[j], color=color, font=fontsize)

    def w2col(self, startrow, jcol, v, font=None):
        i=0
        for x in v:
            self.cell(startrow+i, jcol, value=v[i],font=font)
            i+=1

    # Read into the array from the row
    def rrow(self, row, col, n):
        a=[]
        for j in range(0, n):
            a.append(self.cell(row, col+j))
        return a
    def rcolumn(self, row, col, n):
        a=[]
        for j in range(0, n):
            a.append(self.cell(row+j, col))
        return a


    # Checks if this is a table row.  Returns the number of the table if it is.
    # If it is not a table row, it returns 0.
    def tablerow(self, row):
        s=self.cell(row=row, column=3)
        if s==None or type(s)!=str:
            return 0
        if s=="Table":
            return (self.cell(row=row, column=4))
        else:
            return 0

    def setcellmenu(self, row, column, choices=None):
        from openpyxl.worksheet.datavalidation import DataValidation
        if choices!=None:
            self.ws.menudv = DataValidation(type="list", formula1=choices, allow_blank=False, showDropDown = True)
            self.ws.menudv.prompt=choices
            self.ws.menudv.error="Valid entries "+choices
        self.ws.menudv.add(self.ws.cell(row=row, column=column))
        self.ws.add_data_validation(self.ws.menudv)

    def setcellcolor(self, row, column, color):
        from openpyxl.styles import colors, Font
        cello=self.ws.cell(row=row, column=column)
        cello.font=Font(color=colortable[color])
    def setentrycolor(self, rowname, column, color):
        row=self.findrow(rowname)
        self.setcellcolor(row, column, color)

    def setcellfill(self, row, column, fillcolor):
        from openpyxl.styles import colors, PatternFill
        cello=self.ws.cell(row=row, column=column)
        cello.fill=PatternFill("solid", fgColor=fillcolor)

    wantedtype=None
    exceptions=0.0
    def celldefault(self, t=float, x=0.0):
        self.wantedtype=t
        self.exceptions=x

    def force(self,v):
        if self.wantedtype==None:
            return v
        try:
            if type(v)==self.wantedtype:
                return v
            elif type(v)==str:
                try:
                    z=float(v)
                except:
                    z=self.exceptions
                return z
        except:
            if type(v)==self.wantedtype:
                return v
            return self.exceptions

    def cell(self, row, column, value=None, color="BLACK", font=None, align=None):
        from openpyxl.styles import colors, Font
        cello=self.ws.cell(row=row, column=column)
        if value==None:
            s=cello.value
            if type(s)==str:
                if s[0]=='\ufeff':  # This character sometimes creeps in for files coming from MAC
                    s=s[1:]
                try:
                    z=float(s)
                except:
                    z=s
                s=z
            # s=cello.value.replace(u'\ufeff', '', 1)
            if font!=None:
                cello.font=Font(size=font)
            if type(s)==str:
                if s=="":     return 0
                if s[0]=='=': return self.force(self.ws[s[1:]].value)
                return self.force(s)
            else:
                return self.force(s)
        else:
            cello.value=value
            cello.font=Font(color=colortable[color])
            if font!=None:
                cello.font=Font(size=font,color=colortable[color])
            if align!=None:
                cello.alignment=Alignment(horizontal=align)
            self.lastrow=row
            return value

    def param(self, rowname, column, value=None, color="BLACK"):
        from openpyxl.styles import colors, Font
        row=self.findrow(rowname)
        if value==None:
             return self.cell(row=row, column=column)
        else:
            cello=self.ws.cell(row=row, column=column)
            cello.value=value
            cello.font=Font(color=colortable[color])
            return value

    def colwidth(self, column, width=None):
        from openpyxl.utils import get_column_letter
        c=get_column_letter(column)
        if width==None:
            return (self.ws.column_dimensions[c].width)
        else:
            self.ws.column_dimensions[c].width=width+0.64  # 2/2019: I need 0.64

    def parsevar(self, v, x2):
        if v[0]=='(':
            v=v[1:].split(",")
            vmin=float(v[0])
            vmax=float(v[1].split(")")[0])
            return("DATA RANGE", vmin, vmax, 0)
        elif v[0]=='"':  # pull down menu
#            choices = "'"+v+"'"
            return ("MENU", v, 0, 0)
        v=v.lstrip()
        z=v.split("[",1)
        rowname=z[0]
        s=z[1][:-1]
        column=int(s)
#        k=self.findrow(rowname)
        k=x2.findrow(rowname)
        return ("SINGLE", self.cell(row=k, column=column), k, column)

    def dvrange(self, row, column, vmin, vmax):
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(type="decimal", operator="between", formula1=vmin, formula2=vmax, allow_blank=False)
        choice="a value between %.2f and %.2f"%(vmin, vmax)
        dv.prompt="Enter "+choice
        dv.error="Valid entries are "+choice
        c=get_column_letter(column)
        dv.ranges.append(c+'%d'%row)
        self.ws.add_data_validation(dv)

    def pulldownmenu(self, row, column, choices):
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(type="list", formula1=choices, allow_blank=False, showDropDown = True)
        dv.prompt=choices
        dv.error="Valid entries "+choices
        c=get_column_letter(column)
        a=c+'%d'%row
#        dv.ranges.append(a)
        dv.ranges.add(a)
        self.ws.add_data_validation(dv)


    def hidesheet(self):
        self.ws.sheet_state='hidden'
        self.hidecolumns()

    def hidecolumns(self):
        for j in range(0, self.ws.max_column):
            c=get_column_letter(j+1)
            self.ws.column_dimensions[c].width=0

    def openfile(self, filename, sheetname="", data_only=False):
        try:
            self.wb=load_workbook(filename=filename, data_only=data_only)
        except:
            # sys.exit("%s cannot be found"%filename)
            return False
        if sheetname=="":
            self.ws=self.wb.active
        else:
            self.ws=self.wb[sheetname]
        return True
    def savefile(self, filename):
        try:
            self.Xcelfilename=filename
            self.wb.save(filename)
            return True
        except:
            print("Failed to save "+filename)
            return False
    def addimage(self, picfile, row, column=1):
        img=openpyxl.drawing.image.Image(picfile)
        img.anchor(self.ws.cell(row=row, column=column)) # upper right corner of the image
        self.ws.add_image(img)

    def addfunc(self, function_name):
        self.fdic[function_name]


defaultcellcolor=43
thin_border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
from openpyxl.utils import get_column_letter
def copycellstyles(cell, cello, row, col):
    try:
        color_index=cell.fill.start_color.index
        if type(color_index)==int:
        # Protection against stupid Excel feature.  The colors selected from
        # "theme colors" menu in Excel are integers. There is no color match
        # between my "Excel theme color" indices and what openpyxl thinks they are.
        # I think it is always a string when you pick the color from the color wheel
        # Therefore, I will not accept integers and force selection from the wheel.
            # import sys
            # sys.exit("'%s'(%d,%d) : You must select from the color wheel not from 'standard colors'"%(cell.value,row,col))
            colored=0
        else:
            colored=int(color_index,16)>0
#
    except ValueError:
        import sys
        sys.exit("COPYCELLSTYLES error.  color_index='%s'"%color_index)
    cello.value=cell.value
    cello.number_format=cell.number_format
    cello.font=cell.font.copy(bold=cell.font.bold, size=cell.font.size)
    if cell.border.left.style!=None:
        cello.border=thin_border
    if colored:
        cello.fill=PatternFill(fill_type='solid', start_color=color_index)
        cello.border=thin_border
        cello.value=None
        cello.protection=Protection(locked=False, hidden=False)
    else:
        cello.protection=Protection(locked=True, hidden=False)

def copystyles(x1, r1, c1, x2, r2, c2):
    cell=x1.ws.cell(row=r1, column=c1)
    cello=x2.ws.cell(row=r2, column=c2)
    copycellstyles(cell, cello, r1, c1)

from openpyxl.utils import get_column_letter
def copyrow(x1, r1, x2, r2):
    for j in range(0, x1.ws.max_column):
        cell =x1.ws.cell(row=r1, column=j+1)
        cello=x2.ws.cell(row=r2, column=j+1)
        copycellstyles(cell,cello, r1, j+1)
        s=cell.value
        cello.value=s  # I added this on 9 July 2019
        if s!=None and type(s)==str:
#            print(s+" ", end="")
            if s[0]=='`':  # This was "$" before Feb 2019
                (value, a, b, c)=x1.parsevar(s[1:], x2)
                if value=="DATA RANGE":
                    x2.dvrange(r2, j+1, a, b)
                    cello.value=""  # I added this in Feb 2019
                elif value=="MENU":
                    x2.pulldownmenu(r2, j+1, a)  # a=choices for pulldown menu
                else:
                    cell_address="%s%d"%(get_column_letter(c),b)
                    cello.value="="+cell_address
            elif s[0]=="=":
                cello.value=s

def copycolumnwidths(x1, x2):
    for j in range(0, x1.ws.max_column):
        c=get_column_letter(j+1)
        column_width=x1.ws.column_dimensions[c].width
        if column_width!=None:
            x2.ws.column_dimensions[c].width=column_width

def cellalpharef(row, column):
    c=get_column_letter(column)
    return (c+"%d"%row)


def copyfile(xf1,xf2):
    for i in range(1, xf1.maxrow()+1):
        try:
            copyrow(xf1, i, xf2, i)
        except:
            print("\ncopyrow(%d) error"%i)
    copycolumnwidths(xf1, xf2)

def excelfile(filename, sheetname=""):
    xf1=Xcel(filename, sheetname=sheetname)
    xf2=Xcel()
    copyfile(xf1,xf2)
    # if tofile=="":
    #     tofile=filename
    # xf2.savefile(tofile)
    return xf2

def readcolors():
    x=Xcel(filename="colors.xlsx")
    y=Xcel()
    for i in range(0, x.maxrow()):
        color=x.ws.cell(row=i+1, column=1)
        name=x.ws.cell(row=i+1, column=2).value
        color_index=color.fill.start_color.index
        is_int=type(color_index)==int
        print("Row %d: %s --> '%s'  Is it integer ? %s"%(i+1, name, color_index, is_int))
        copyrow(x, i+1, y, i+1)
        y.cell(i+1, 5, value="'%s'"%color_index)
    y.savefile("coloroutput.xlsx")
#readcolors()


if TESTMODULE=="readcolor":
    inputfile="template"
    Xf=Xcel(inputfile+".xlsx")
    Xf2  =Xcel(sheetname="data")
    for i in range (1, Xf.maxrow()):
        copyrow(Xf, i, Xf2, i)
    copycolumnwidths(Xf, Xf2)
    Xf2.savefile(inputfile+"x.xlsx")
if TESTMODULE=="fontcolor":  # Test column widths and styles
    Xf=Xcel()
    Xf.cell(1,1,value="RED",font=8, color="RED")
    Xf.cell(2,1,value="BLACK",font=8, color="BLACK")
    Xf.savefile("fontcolor.xlsx")
if TESTMODULE=="w2row":
    Xf=Xcel()
    Xf.cell(1,1,value="W2ROW")
    Xf.cell(2,1,value="1,2,3,4")
    Xf.w2row(2,2,[1,2,3,4])
    Xf.cell(3,1,value="['one', 'two', 1.00, 2]")
    Xf.w2row(3,2,['one', 'two', 1.00, 2])
    Xf.savefile("w2row.xlsx")
if TESTMODULE=="vit":
    Xm=Xcel("mats.xlsx")
    ro=Xm.vit(3,"STEEL", "Density")
    print(ro)
